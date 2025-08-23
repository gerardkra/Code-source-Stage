# ============================================
# app.py - Démo Streamlit Recommandation Produits
# ============================================
import altair as alt
import io
from openpyxl import Workbook

import hashlib

import streamlit as st
import numpy as np
import pandas as pd
import re, string
from collections import Counter, defaultdict
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.neighbors import NearestNeighbors
from scipy.sparse import hstack, csr_matrix
import joblib
import warnings
warnings.filterwarnings("ignore")

# Charger les données MovieLens
df = pd.read_csv(
    r'C:\Users\hp\OneDrive - UNIVERSITE DES LAGUNES\Downloads (2)\STAGE_M2\Doc_Stage\data_info_client_clean.csv',
    low_memory=False
)

# ============================================
# 1) Artefacts
# ============================================
word2id = joblib.load("word2id.joblib")
embedding_matrix = np.load("embedding_matrix.npy")
embedding_size = embedding_matrix.shape[1]
id2word = {v: k for k, v in word2id.items()}

# ============================================
# 2) Fonctions utilitaires
# ============================================
def clean_product_name(name):
    if name is None:
        return ""
    name = str(name).lower()
    name = name.translate(str.maketrans('', '', string.punctuation))
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def tokenize_vecteur_prod(vprod_str):
    if pd.isna(vprod_str):
        return []
    return [clean_product_name(p) for p in str(vprod_str).split('|') if p]

def client_vector_from_tokens(tokens, embedding_matrix, word2id, embedding_size):
    idx = [word2id[t] for t in tokens if t in word2id]
    if not idx:
        return np.zeros(embedding_size, dtype='float32')
    return embedding_matrix[idx].mean(axis=0).astype('float32')

def is_zero_vector(v, tol=1e-8):
    v = np.asarray(v)
    return np.all(np.abs(v) < tol)

def product_vec(token):
    if token in word2id:
        return embedding_matrix[word2id[token]]
    return None

# ============================================
# 3) Charger les données
# ============================================
# ⚠️ Mets ici ton fichier CSV clients (change le nom si nécessaire)
df = df.copy()

if "VECTEUR_PROD" not in df.columns:
    raise ValueError("La colonne 'VECTEUR_PROD' est introuvable dans df.")

df["tokens"] = df["VECTEUR_PROD"].fillna("").apply(tokenize_vecteur_prod)
df["tokens_known"] = df["tokens"].apply(lambda toks: [t for t in toks if t in word2id])
df["vecteur_client"] = df["tokens_known"].apply(
    lambda toks: client_vector_from_tokens(toks, embedding_matrix, word2id, embedding_size)
)
embeddings = np.vstack(df["vecteur_client"].to_numpy())

# ============================================
# 4) IDF anti-popularité
# ============================================
def build_idf(df, word2id):
    N = len(df)
    df_counts = Counter()
    for toks in df["tokens_known"]:
        df_counts.update(set(toks))
    idf_raw = {t: np.log((N + 1) / (df_counts.get(t, 0) + 1)) + 1.0 for t in word2id.keys()}
    max_idf = max(idf_raw.values()) if idf_raw else 1.0
    idf_norm = {t: (v / max_idf) for t, v in idf_raw.items()}
    return idf_norm, df_counts

idf_norm, df_counts = build_idf(df, word2id)

# ============================================
# 5) Encodage PROFIL + kNN
# ============================================
cat_cols = ["SEXE","TITUPRINC","SITUATIONMATRIMONAL","SEGMENT","PROFESSION","CIVILITE"]
num_cols = ["AGE","ANCIENNETE_JOURS"]

for c in cat_cols:
    if c not in df.columns:
        df[c] = "NA"
for c in num_cols:
    if c not in df.columns:
        df[c] = 0

df_cat = df[cat_cols].fillna("NA").astype(str)
df_num = df[num_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)

try:
    enc = OneHotEncoder(handle_unknown="ignore", sparse_output=True)
except TypeError:
    enc = OneHotEncoder(handle_unknown="ignore", sparse=True)

X_cat = enc.fit_transform(df_cat)
scaler = StandardScaler()
X_num = scaler.fit_transform(df_num.values.astype(float))
X = hstack([X_cat, csr_matrix(X_num)])

n_neighbors_profile = 100
knn_profile = NearestNeighbors(n_neighbors=n_neighbors_profile, metric="cosine")
knn_profile.fit(X)

def encode_profile_row(row_dict):
    row_cat = {c: str(row_dict.get(c, "NA")) if pd.notna(row_dict.get(c, None)) else "NA" for c in cat_cols}
    df_row_cat = pd.DataFrame([row_cat], columns=cat_cols)
    Xc = enc.transform(df_row_cat)

    row_num = {}
    for c in num_cols:
        v = row_dict.get(c, 0)
        try:
            v = float(v)
        except:
            v = 0.0
        row_num[c] = v
    df_row_num = pd.DataFrame([row_num], columns=num_cols)
    Xn = scaler.transform(df_row_num.values.astype(float))

    return hstack([Xc, csr_matrix(Xn)])

# ============================================
# 6) Diversification MMR
# ============================================
def mmr_rerank(candidates, top_n=5, lambda_mmr=0.7):
    if not candidates:
        return []
    selected, remaining = [], set(candidates.keys())
    vecs = {t: product_vec(t) for t in candidates.keys()}

    def sim_prod(a, b):
        va, vb = vecs.get(a), vecs.get(b)
        if va is None or vb is None:
            return 0.0
        denom = (np.linalg.norm(va)+1e-9) * (np.linalg.norm(vb)+1e-9)
        return float(np.dot(va, vb) / denom)

    while remaining and len(selected) < top_n:
        best_t, best_score = None, -1e9
        for t in list(remaining):
            relevance = float(candidates[t])
            redundancy = max([sim_prod(t, s) for s in selected], default=0.0)
            mmr_score = lambda_mmr * relevance - (1 - lambda_mmr) * redundancy
            if mmr_score > best_score:
                best_score, best_t = mmr_score, t
        selected.append(best_t)
        remaining.remove(best_t)
    return [(t, float(candidates[t])) for t in selected]

# ============================================
# 7) Voisins effectifs
# ============================================
def effective_neighbors(weights, idxs, tau=0.8):
    order = np.argsort(weights)[::-1]
    w_sorted = weights[order]
    i_sorted = idxs[order]
    total = w_sorted.sum() + 1e-9
    cumsum = np.cumsum(w_sorted)
    k_eff = int(np.searchsorted(cumsum, tau * total, side="left")) + 1
    return w_sorted[:k_eff], i_sorted[:k_eff]

# ============================================
# 8) Recommandation Hybride Optimisée
# ============================================
def recommend_products_hybride_optim(
    client_index, df, embeddings,
    top_n=5,
    min_similarity=0.10,
    use_mmr=True,
    lambda_mmr=0.7,
    idf_alpha=0.30,
    tau_neighbors=0.80,
    score_mode="relative"
):
    owned = set(df.iloc[client_index]["tokens_known"])
    cvec = embeddings[client_index]

    if not is_zero_vector(cvec) and len(owned) > 0:
        sims = cosine_similarity(cvec.reshape(1, -1), embeddings).ravel()
        candidate_scores = {}
        for idx in np.argsort(sims)[::-1][1:]:
            if sims[idx] < min_similarity:
                break
            for t in [t for t in df.iloc[idx]["tokens_known"] if t not in owned]:
                candidate_scores[t] = candidate_scores.get(t, 0.0) + float(sims[idx]) * 0.5

        if not candidate_scores:
            return []

        max_s = max(candidate_scores.values())
        rel_scores = {t: (s / (max_s + 1e-12)) for t, s in candidate_scores.items()}

        if use_mmr:
            mmr = mmr_rerank(rel_scores, top_n=top_n, lambda_mmr=lambda_mmr)
            return [("produit", t, s) for t, s in mmr if s >= 0.70]
        else:
            recs = sorted(rel_scores.items(), key=lambda x: -x[1])[:top_n]
            return [("produit", t, float(s)) for t, s in recs if s >= 0.70]

    row = df.iloc[client_index].to_dict()
    X_row = encode_profile_row(row)
    dists, idxs = knn_profile.kneighbors(X_row, return_distance=True)
    dists, idxs = dists[0], idxs[0]

    weights = np.maximum(1.0 - dists, 0.0)
    w_eff, i_eff = effective_neighbors(weights, idxs, tau=tau_neighbors)
    total_w = float(w_eff.sum()) + 1e-9

    base_scores = defaultdict(float)
    cover = defaultdict(int)

    for w, j in zip(w_eff, i_eff):
        if j == client_index:
            continue
        for t in df.iloc[j]["tokens_known"]:
            if t not in owned:
                base_scores[t] += float(w)
                cover[t] += 1

    if not base_scores:
        return []

    for t in list(base_scores.keys()):
        base_scores[t] = base_scores[t] / total_w

    if idf_alpha > 0:
        for t in list(base_scores.keys()):
            base_scores[t] *= (1.0 - idf_alpha) + idf_alpha * idf_norm.get(t, 0.0)

    if score_mode == "relative":
        m = max(base_scores.values())
        final_scores = {t: base_scores[t] / (m + 1e-12) for t in base_scores}
    else:
        final_scores = base_scores

    if use_mmr:
        mmr = mmr_rerank(final_scores, top_n=top_n, lambda_mmr=lambda_mmr)
        return [("lookalike", t, s, cover[t]) for t, s in mmr if s >= 0.70]
    else:
        recs = sorted(final_scores.items(), key=lambda x: -x[1])[:top_n]
        return [("lookalike", t, float(s), cover[t]) for t, s in recs if s >= 0.70]

# ============================================
# 9) Interface Streamlit

# Fonction pour nettoyer le texte
# ====================================================
# ====================================================
# 0) Imports
# ====================================================

def build_client_profile(row):
    """
    Construit un résumé textuel simple du profil client
    à partir des colonnes principales du dataframe.
    """
    parts = []

    # Exemples de champs que tu peux inclure
    if "SEXE" in row and pd.notna(row["SEXE"]):
        parts.append(f"Sexe : {row['SEXE']}")
    if "AGE" in row and pd.notna(row["AGE"]):
        parts.append(f"Âge : {row['AGE']} ans")
    if "SEGMENT" in row and pd.notna(row["SEGMENT"]):
        parts.append(f"Segment : {row['SEGMENT']}")
    if "PROFESSION" in row and pd.notna(row["PROFESSION"]):
        parts.append(f"Profession : {row['PROFESSION']}")
    if "SITUATIONMATRIMONAL" in row and pd.notna(row["SITUATIONMATRIMONAL"]):
        parts.append(f"Situation : {row['SITUATIONMATRIMONAL']}")

    return " | ".join(parts) if parts else "Profil client non disponible"


# ====================================================
# 1) Fonction de hachage
# ====================================================
def hash_sha512(texte):
    """Retourne le hash SHA-512 d'une valeur en string"""
    if pd.isna(texte):
        return None
    return hashlib.sha512(str(texte).encode("utf-8")).hexdigest()

# ====================================================
# 2) Recherche par IDP ou CLIENT
# ====================================================
def match_identifiant(df, saisie_idp, saisie_client):
    """
    Retourne l'index du client trouvé via IDP ou CLIENT.
    Accepte en entrée une valeur brute ou déjà hachée.
    """
    if saisie_idp:
        saisie_idp_hash = hash_sha512(saisie_idp)
        for i, row in df.iterrows():
            if str(row["IDP"]) in [saisie_idp, saisie_idp_hash]:
                return i

    if not saisie_idp and saisie_client:
        saisie_client_hash = hash_sha512(saisie_client)
        for i, row in df.iterrows():
            if str(row["CLIENT"]) in [saisie_client, saisie_client_hash]:
                return i

    return None

# ====================================================
# 3) Interface Streamlit
# ====================================================
st.set_page_config(page_title="Moteur de Recommandation Produits", layout="wide")
st.title("✨ Démo - Moteur de Recommandation Produits")

# --- Sidebar ---
with st.sidebar:
    st.header("⚙️ Paramètres")

    # --- Liste déroulante IDP ---
    idp_list = df["IDP"].dropna().unique().tolist()
    idp_choice = st.selectbox("🔽 Choisir un IDP existant", [""] + idp_list, index=0)

    # --- Saisie manuelle IDP ---
    saisie_idp_input = st.text_input("✏️ Ou saisir un IDP (réel ou hash)")
    saisie_idp = saisie_idp_input if saisie_idp_input.strip() else idp_choice

    # --- Liste déroulante CLIENT ---
    client_list = df["CLIENT"].dropna().unique().tolist()
    client_choice = st.selectbox("🔽 Choisir un CLIENT existant", [""] + client_list, index=0)

    # --- Saisie manuelle CLIENT ---
    saisie_client_input = st.text_input("✏️ Ou saisir un CLIENT (réel ou hash)")
    saisie_client = saisie_client_input if saisie_client_input.strip() else client_choice

    # --- Autres paramètres ---
    top_n = st.slider("Nombre de recommandations", 1, 10, 5)
    st.markdown("---")
    advanced = st.checkbox("🔧 Mode avancé")

    if advanced:
        min_similarity = st.slider("Min Similarité", 0.0, 1.0, 0.6, 0.05)
        lambda_mmr = st.slider("Lambda MMR", 0.0, 1.0, 0.7, 0.05)
        tau_neighbors = st.slider("Tau voisins", 0.0, 1.0, 0.8, 0.05)
        idf_alpha = st.slider("IDF α", 0.0, 1.0, 0.2, 0.05)
        score_mode = st.radio("Mode score", ["relative", "absolute"], index=0)
    else:
        min_similarity = 0.60
        lambda_mmr = 0.7
        tau_neighbors = 0.80
        idf_alpha = 0.20
        score_mode = "relative"

# ====================================================
# 4) Action principale
# ====================================================
if st.button("🚀 Obtenir les recommandations"):

    client_index = match_identifiant(df, saisie_idp, saisie_client)

    if client_index is None:
        if saisie_idp:
            st.error("❌ Aucun client trouvé avec cet IDP.")
        else:
            st.warning("ℹ️ Ce client n'a pas d'IDP. Veuillez saisir sa racine CLIENT.")
        st.stop()

    # --- Génération des recommandations ---
    recs = recommend_products_hybride_optim(
        client_index=client_index,
        df=df,
        embeddings=embeddings,
        top_n=top_n,
        min_similarity=min_similarity,
        use_mmr=True,
        lambda_mmr=lambda_mmr,
        idf_alpha=idf_alpha,
        tau_neighbors=tau_neighbors,
        score_mode=score_mode
    )

    row = df.iloc[client_index]
    profile_text = build_client_profile(row)
    owned = df.iloc[client_index]["tokens_known"]

    data = []
    for r in recs:
        if r[0] == "produit":
            _, prod, score = r
            data.append([prod, score])
        else:
            _, prod, score, _ = r
            data.append([prod, score])

    df_recs = pd.DataFrame(data, columns=["Produit", "Score"])

    # ====================================================
    # 5) Onglets
    # ====================================================
    tab1, tab2, tab3 = st.tabs(["👤 Profil client", "📦 Produits possédés", "✨ Recommandations"])

    with tab1:
        st.write("**Profil généré :**")
        st.info(profile_text)

        colonnes_utiles = [
            "DATNAIS", "DATE_EER", "SEXE", "TITUPRINC",
            "SITUATIONMATRIMONAL", "SEGMENT", "PROFESSION", "LIBELLE_AGENT_ECO",
            "CIVILITE", "AGE", "ANCIENNETE_JOURS"
        ]
        st.write("Données brutes :")
        st.dataframe(df.loc[[client_index], colonnes_utiles].T, use_container_width=True)

    with tab2:
        st.write("Produits déjà possédés :")
        st.success(owned)

    with tab3:
        st.subheader("Tableau des recommandations")
        st.dataframe(df_recs.style.format({"Score": "{:.4f}"}), use_container_width=True)

        st.subheader("Graphique des scores")
        chart = alt.Chart(df_recs).mark_bar().encode(
            x="Score:Q",
            y=alt.Y("Produit:N", sort='-x'),
            tooltip=["Produit", "Score"]
        ).properties(height=400)
        st.altair_chart(chart, use_container_width=True)

    # ====================================================
    # 6) Export CSV
    # ====================================================
    csv_buffer = io.StringIO()
    df_recs.to_csv(csv_buffer, index=False)
    csv_data = csv_buffer.getvalue()

    st.download_button(
        label="📥 Télécharger les recommandations (.csv)",
        data=csv_data,
        file_name=f"recommandations_client_{client_index}.csv",
        mime="text/csv"
    )

    # ====================================================
    # 7) Export XLSX
    # ====================================================
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Profil client"
    ws1.append(["Profil"])
    ws1.append([profile_text])

    ws2 = wb.create_sheet("Produits possédés")
    ws2.append(["Produits"])
    if isinstance(owned, (list, set)):
        for prod in owned:
            ws2.append([prod])
    else:
        ws2.append([owned])

    ws3 = wb.create_sheet("Recommandations")
    ws3.append(df_recs.columns.tolist())
    for row in df_recs.itertuples(index=False):
        ws3.append(list(row))

    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)

    st.download_button(
        label="📥 Télécharger le rapport complet (.xlsx)",
        data=xlsx_buffer,
        file_name=f"rapport_client_{client_index}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
